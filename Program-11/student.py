from openpyxl import load_workbook

# Excel file path
file_path = r"D:\MCA\SEM-2\Software Testing\Lab programs\Program-11\students.xlsx"

# Load workbook
wb = load_workbook(file_path)

# Select active sheet
sheet = wb.active

# Update marks for Roll No 103
for row in sheet.iter_rows(min_row=2):
    if row[0].value == 103:
        row[2].value = 99
        break

# Save workbook
wb.save(file_path)

print("Student record updated successfully.")